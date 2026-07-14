import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

file_path = "Projet_ACP.xlsx"
output_path = "Resultats_ACP_Pays_V2.xlsx"

df = pd.read_excel(file_path, sheet_name="Country-data")
numeric_df = df.select_dtypes(include=[np.number])
variables = numeric_df.columns.tolist()
data = numeric_df.values.astype(float)
pays = df['country'].tolist()

# Standardisation
mean = np.mean(data, axis=0)
std = np.std(data, axis=0, ddof=1)
scaled = (data - mean) / std

# Matrice corrélation + régularisation
corr = np.corrcoef(scaled.T)
corr = corr + np.eye(len(variables)) * 1e-8

# ACP
eigenvalues, eigenvectors = np.linalg.eigh(corr)
idx = np.argsort(eigenvalues)[::-1]
eigenvalues = eigenvalues[idx]
eigenvectors = eigenvectors[:, idx]
variance_ratio = eigenvalues / eigenvalues.sum()
cum_variance = np.cumsum(variance_ratio)

# Création Excel
wb = Workbook()
ws = wb.active
ws.title = "ACP_Pays"

row = 1

# Titre
ws.cell(row, 1, "Analyse en Composantes Principales - Données Pays").font = Font(bold=True, size=14)
row += 2

# Moyenne et Ecart-type
ws.cell(row, 1, "Moyenne").font = Font(bold=True)
for i, v in enumerate(mean, 2):
    ws.cell(row, i, round(v, 4))
row += 1
ws.cell(row, 1, "Ecart-type").font = Font(bold=True)
for i, v in enumerate(std, 2):
    ws.cell(row, i, round(v, 4))
row += 3

# Matrice de corrélation
ws.cell(row, 1, "Matrice de corrélation").font = Font(bold=True)
row += 1
for i, var in enumerate(variables):
    ws.cell(row, i+2, var)
row += 1
for i, var in enumerate(variables):
    ws.cell(row, 1, var)
    for j in range(len(variables)):
        ws.cell(row, j+2, round(corr[i,j], 4))
    row += 1
row += 2

# Valeurs propres
ws.cell(row, 1, "Valeurs propres").font = Font(bold=True)
row += 1
ws.cell(row, 1, "Composante")
ws.cell(row, 2, "Valeur propre")
ws.cell(row, 3, "% variance")
ws.cell(row, 4, "% cumulé")
row += 1
for i in range(min(5, len(eigenvalues))):
    ws.cell(row, 1, f"PC{i+1}")
    ws.cell(row, 2, round(eigenvalues[i], 4))
    ws.cell(row, 3, f"{round(variance_ratio[i]*100, 2)}%")
    ws.cell(row, 4, f"{round(cum_variance[i]*100, 2)}%")
    row += 1
row += 2

# Vecteurs propres
ws.cell(row, 1, "Vecteurs propres").font = Font(bold=True)
row += 1
ws.cell(row, 1, "Variable")
for j in range(4):
    ws.cell(row, j+2, f"PC{j+1}")
row += 1
for i, var in enumerate(variables):
    ws.cell(row, 1, var)
    for j in range(4):
        ws.cell(row, j+2, round(eigenvectors[i, j], 4))
    row += 1
row += 2

# Coordonnées des pays
ws.cell(row, 1, "Coordonnées des individus").font = Font(bold=True)
row += 1
ws.cell(row, 1, "Pays")
for j in range(4):
    ws.cell(row, j+2, f"PC{j+1}")
row += 1

for i in range(len(pays)):
    ws.cell(row, 1, pays[i])
    for j in range(4):
        ws.cell(row, j+2, round(scaled[i] @ eigenvectors[:, j], 3))
    row += 1

wb.save(output_path)
print(f"✅ Version améliorée créée : {output_path}")


import matplotlib.pyplot as plt
import seaborn as sns

# Graphique 1 : Plan factoriel PC1 vs PC2
plt.figure(figsize=(12, 8))
scores = scaled @ eigenvectors

plt.scatter(scores[:, 0], scores[:, 1], alpha=0.7)

# Afficher quelques pays intéressants
for i in range(len(pays)):
    if abs(scores[i,0]) > 2 or abs(scores[i,1]) > 1.5:   # Les plus extrêmes
        plt.annotate(pays[i], (scores[i,0], scores[i,1]), fontsize=8)

plt.xlabel('PC1 - Développement (45.95%)')
plt.ylabel('PC2 - Ouverture Commerciale (17.18%)')
plt.title('Plan Factoriel PC1 vs PC2 - Données Pays')
plt.grid(True)
plt.axhline(0, color='gray', linestyle='--')
plt.axvline(0, color='gray', linestyle='--')
plt.savefig('Plan_Factoriel_PC1_PC2.png', dpi=300, bbox_inches='tight')
print("Graphique Plan Factoriel sauvegardé : Plan_Factoriel_PC1_PC2.png")

# Graphique 2 : Cercle des corrélations
plt.figure(figsize=(10, 10))
for i, var in enumerate(variables):
    plt.arrow(0, 0, eigenvectors[i,0], eigenvectors[i,1], 
              head_width=0.05, head_length=0.05, fc='blue', ec='blue')
    plt.text(eigenvectors[i,0]*1.1, eigenvectors[i,1]*1.1, var, color='red')

plt.xlim(-1.1, 1.1)
plt.ylim(-1.1, 1.1)
plt.xlabel('PC1')
plt.ylabel('PC2')
plt.title('Cercle des Corrélations (PC1-PC2)')
plt.grid(True)
circle = plt.Circle((0, 0), 1, color='gray', fill=False, linestyle='--')
plt.gca().add_artist(circle)
plt.axhline(0, color='gray', linestyle='--')
plt.axvline(0, color='gray', linestyle='--')
plt.savefig('Cercle_Correlations.png', dpi=300, bbox_inches='tight')
print("Graphique Cercle des corrélations sauvegardé : Cercle_Correlations.png")